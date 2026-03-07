
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from shared.logger import get_logger
from shared.utils.export_media import prepare_cited_assets, strip_export_citation_tags
from shared.utils.export_text import build_export_blocks, format_structured_text, inline_tokens_to_text, strip_markdown

logger = get_logger(__name__)

_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F4F6F", end_color="2F4F6F", fill_type="solid")
_EVEN_ROW_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

_CONF_FILLS = {
    "high": PatternFill(start_color="E6F5EC", end_color="E6F5EC", fill_type="solid"),
    "medium": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    "low": PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid"),
}
_CONF_FONTS = {
    "high": Font(bold=True, color="1B7A3D"),
    "medium": Font(bold=True, color="B8860B"),
    "low": Font(bold=True, color="CC3333"),
}

def _configure_sheet(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "$1:$1"


def _sheet_text(value: str) -> str:
    text = strip_export_citation_tags(strip_markdown(str(value or "")))
    return format_structured_text(text)

def _set_cell(ws: Worksheet, row: int, col: int, value: str,
              font: Font | None = None, fill: PatternFill | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = _WRAP_ALIGNMENT
    cell.border = _THIN_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill

def _auto_column_widths(ws: Worksheet, min_width: int = 12, max_width: int = 65) -> None:
    for col_cells in ws.columns:
        if not col_cells or col_cells[0] is None:
            continue
        first = col_cells[0]
        if not isinstance(first, Cell):
            continue
        longest = max(
            (len(str(c.value or "")) for c in col_cells if c is not None),
            default=min_width,
        )
        ws.column_dimensions[first.column_letter].width = max(
            min_width, min(longest + 3, max_width)
        )

def _set_header_row(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_ALIGNMENT
        cell.border = _THIN_BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"

def _apply_alternating_rows(ws: Worksheet, start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        if row[0].row % 2 == 0:
            for cell in row:
                if cell.fill == PatternFill():
                    cell.fill = _EVEN_ROW_FILL

def _populate_comparison_rows(ws: Worksheet, paper_data: list[dict]) -> None:
    for row_idx, paper in enumerate(paper_data, start=2):
        _set_cell(ws, row_idx, 1, str(paper.get("title", "")))
        _set_cell(ws, row_idx, 2, str(paper.get("authors", "")))
        ws.cell(row=row_idx, column=3, value=paper.get("year", "")).border = _THIN_BORDER
        _set_cell(ws, row_idx, 4, str(paper.get("abstract", "") or ""))
        _set_cell(ws, row_idx, 5, str(paper.get("problem", "") or ""))
        _set_cell(ws, row_idx, 6, str(paper.get("method", "") or ""))
        _set_cell(ws, row_idx, 7, str(paper.get("results", "") or ""))
        _set_cell(ws, row_idx, 8, str(paper.get("keywords", "") or ""))


def _split_for_sheet(value: str, chunk_size: int = 320) -> list[str]:
    normalized = _sheet_text(value)
    if len(normalized) <= chunk_size:
        return [normalized]

    chunks: list[str] = []
    remaining = normalized
    while remaining:
        if len(remaining) <= chunk_size:
            chunks.append(remaining)
            break
        split_at = remaining.rfind(" ", 0, chunk_size)
        if split_at <= 0:
            split_at = chunk_size
        chunks.append(remaining[:split_at].strip())
        remaining = remaining[split_at:].strip()
    return [chunk for chunk in chunks if chunk]


def _populate_chat_blocks_sheet(ws: Worksheet, messages: list[dict]) -> None:
    headers = ["Message", "Role", "Block", "Order", "Content"]
    _set_header_row(ws, headers)
    row_idx = 2
    for message_idx, msg in enumerate(messages, start=1):
        role = str(msg.get("role", "user")).upper()
        for block_idx, block in enumerate(build_export_blocks(str(msg.get("content", ""))), start=1):
            block_label = block.kind.title()
            if block.kind == "table":
                for table_row_idx, table_row in enumerate(block.rows, start=1):
                    pairs = []
                    for cell_idx, cell in enumerate(table_row):
                        label = block.headers[cell_idx] if cell_idx < len(block.headers) else f"Column {cell_idx + 1}"
                        pairs.append(f"{label}: {cell}")
                    _set_cell(ws, row_idx, 1, str(message_idx))
                    _set_cell(ws, row_idx, 2, role)
                    _set_cell(ws, row_idx, 3, f"{block_label} Row")
                    _set_cell(ws, row_idx, 4, f"{block_idx}.{table_row_idx}")
                    _set_cell(ws, row_idx, 5, _sheet_text(" | ".join(pairs)))
                    row_idx += 1
                continue

            content = inline_tokens_to_text(block.tokens) if block.tokens else block.text
            for part_idx, part in enumerate(_split_for_sheet(content), start=1):
                _set_cell(ws, row_idx, 1, str(message_idx))
                _set_cell(ws, row_idx, 2, role)
                _set_cell(ws, row_idx, 3, block_label)
                _set_cell(ws, row_idx, 4, f"{block_idx}.{part_idx}")
                _set_cell(ws, row_idx, 5, part)
                row_idx += 1

    _apply_alternating_rows(ws)
    _auto_column_widths(ws, min_width=10, max_width=42)


def _populate_paper_fields_sheet(ws: Worksheet, paper_data: list[dict]) -> None:
    headers = ["Paper", "Field", "Part", "Content"]
    _set_header_row(ws, headers)
    row_idx = 2
    fields = [
        ("authors", "Authors"),
        ("year", "Year"),
        ("abstract", "Abstract"),
        ("problem", "Problem"),
        ("method", "Method"),
        ("results", "Results"),
        ("keywords", "Keywords"),
    ]
    for paper in paper_data:
        paper_title = str(paper.get("title", ""))
        for key, label in fields:
            value = _sheet_text(str(paper.get(key, "") or ""))
            parts = _split_for_sheet(value, chunk_size=280)
            for part_idx, part in enumerate(parts, start=1):
                _set_cell(ws, row_idx, 1, paper_title)
                _set_cell(ws, row_idx, 2, label)
                _set_cell(ws, row_idx, 3, str(part_idx))
                _set_cell(ws, row_idx, 4, part)
                row_idx += 1

    _apply_alternating_rows(ws)
    _auto_column_widths(ws, min_width=12, max_width=40)

def _populate_citation_rows(ws: Worksheet, citations: list[dict]) -> None:
    for row_idx, cit in enumerate(citations, start=2):
        _set_cell(ws, row_idx, 1, _sheet_text(str(cit.get("claim", ""))))

        confidence = str(cit.get("confidence", "low")).lower()
        conf_cell = ws.cell(row=row_idx, column=2, value=confidence.upper())
        conf_cell.font = _CONF_FONTS.get(confidence, _CONF_FONTS["low"])
        conf_cell.fill = _CONF_FILLS.get(confidence, _CONF_FILLS["low"])
        conf_cell.alignment = _WRAP_ALIGNMENT
        conf_cell.border = _THIN_BORDER

        score = cit.get("score", 0.0)
        score_cell = ws.cell(row=row_idx, column=3, value=score)
        score_cell.number_format = "0%"
        score_cell.border = _THIN_BORDER

        _set_cell(ws, row_idx, 4, _sheet_text(str(cit.get("source_sentence", ""))))
        _set_cell(ws, row_idx, 5, str(cit.get("paragraph_id", "")))


def _populate_comparison_table_sheet(
    ws: Worksheet,
    comparison_table: list[dict],
    paper_data: list[dict],
) -> None:
    paper_titles = [str(paper.get("title", "")) for paper in paper_data]
    headers = ["Dimension", *paper_titles, "Synthesis"]
    _set_header_row(ws, headers)

    for row_idx, row in enumerate(comparison_table, start=2):
        _set_cell(ws, row_idx, 1, str(row.get("dimension", "")))
        for col_idx, cell in enumerate(row.get("cells", []), start=2):
            value = _sheet_text(str(cell.get("content", "")))
            _set_cell(ws, row_idx, col_idx, value)
        _set_cell(ws, row_idx, len(headers), _sheet_text(str(row.get("synthesis", ""))))

    _apply_alternating_rows(ws)
    _auto_column_widths(ws, min_width=18, max_width=36)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22 if col_idx == 1 else 30


def _populate_cited_media_sheet(ws: Worksheet, cited_assets: list[dict]) -> None:
    headers = ["Citation", "Type", "Paper", "Page", "Section", "Content", "Image Path"]
    _set_header_row(ws, headers)

    for row_idx, asset in enumerate(cited_assets, start=2):
        _set_cell(ws, row_idx, 1, str(asset.get("citation_id", "")))
        _set_cell(ws, row_idx, 2, str(asset.get("chunk_type", "")))
        _set_cell(ws, row_idx, 3, str(asset.get("paper_title", "")))
        _set_cell(ws, row_idx, 4, str(asset.get("page_number", "") or ""))
        _set_cell(ws, row_idx, 5, str(asset.get("section_title", "") or ""))
        _set_cell(ws, row_idx, 6, _sheet_text(str(asset.get("description") or asset.get("content", ""))))
        relative_image_path = str(asset.get("image_rel_path", "") or "")
        link_cell = ws.cell(row=row_idx, column=7, value=relative_image_path)
        link_cell.alignment = _WRAP_ALIGNMENT
        link_cell.border = _THIN_BORDER
        if relative_image_path:
            link_cell.hyperlink = relative_image_path
            link_cell.style = "Hyperlink"

        image_path = asset.get("image_path")
        if image_path:
            try:
                from openpyxl.drawing.image import Image as XLImage

                if Path(image_path).exists():
                    image = XLImage(str(image_path))
                    image.width = min(image.width, 220)
                    image.height = min(image.height, 160)
                    ws.add_image(image, f"H{row_idx}")
                    ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, 120)
            except Exception:
                pass

    _apply_alternating_rows(ws)
    _auto_column_widths(ws, min_width=14, max_width=42)
    ws.column_dimensions["H"].width = 34

def export_comparison_to_excel(
    paper_data: list[dict],
    output_path: str | Path,
    comparison_content: str = "",
    comparison_table: list[dict] | None = None,
    cited_assets: list[dict] | None = None,
) -> Path:
    output_path = Path(output_path)
    cited_assets = prepare_cited_assets(cited_assets or [], output_path.parent)
    wb = Workbook()

    ws_comp = wb.active
    if ws_comp is None:
        raise ValueError("Failed to create worksheet")
    ws_comp.title = "Comparison"
    _configure_sheet(ws_comp)

    if comparison_table:
        _populate_comparison_table_sheet(ws_comp, comparison_table, paper_data)
    else:
        title_cell = ws_comp.cell(row=1, column=1, value="Paper Comparison Analysis")
        title_cell.font = Font(bold=True, size=14, color="2F4F6F")
        ws_comp.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

        date_cell = ws_comp.cell(
            row=2, column=1,
            value=f"Exported on {datetime.now().strftime('%B %d, %Y')}",
        )
        date_cell.font = Font(italic=True, size=9, color="888888")

        cleaned = format_structured_text(comparison_content) if comparison_content else ""
        cleaned = _sheet_text(cleaned)
        if cleaned:
            row_idx = 4
            for paragraph in cleaned.split("\n\n"):
                paragraph = paragraph.strip()
                if paragraph:
                    _set_cell(ws_comp, row_idx, 1, paragraph)
                    row_idx += 1
        _auto_column_widths(ws_comp)
        ws_comp.column_dimensions["A"].width = 110

    ws_papers = wb.create_sheet(title="Paper Fields")
    _configure_sheet(ws_papers)
    _populate_paper_fields_sheet(ws_papers, paper_data)

    if cited_assets:
        ws_media = wb.create_sheet(title="Cited Media")
        _configure_sheet(ws_media)
        _populate_cited_media_sheet(ws_media, cited_assets)

    wb.save(str(output_path))
    logger.info(f"Exported comparison Excel to {output_path.name}")
    return output_path

def export_citations_to_excel(
    citations: list[dict],
    output_path: str | Path,
    session_title: str = "Chat Export",
    messages: list[dict] | None = None,
    cited_assets: list[dict] | None = None,
) -> Path:
    output_path = Path(output_path)
    cited_assets = prepare_cited_assets(cited_assets or [], output_path.parent)
    wb = Workbook()

    ws_chat = wb.active
    if ws_chat is None:
        raise ValueError("Failed to create worksheet")
    ws_chat.title = "Chat Blocks"
    _configure_sheet(ws_chat)
    if messages:
        _populate_chat_blocks_sheet(ws_chat, messages)
    else:
        _set_header_row(ws_chat, ["Message", "Role", "Block", "Order", "Content"])

    ws_cit = wb.create_sheet(title="Citation Verification")
    _configure_sheet(ws_cit)
    cit_headers = ["Claim", "Confidence", "Score", "Source Sentence", "Paragraph ID"]
    _set_header_row(ws_cit, cit_headers)
    _populate_citation_rows(ws_cit, citations)
    _apply_alternating_rows(ws_cit)
    _auto_column_widths(ws_cit)

    if cited_assets:
        ws_media = wb.create_sheet(title="Cited Media")
        _configure_sheet(ws_media)
        _populate_cited_media_sheet(ws_media, cited_assets)

    wb.save(str(output_path))
    logger.info(f"Exported citations Excel to {output_path.name}")
    return output_path

